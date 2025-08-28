# 中文备注：生成“装备图鉴 Excel 模板（.xlsx）”，包含：
# 1) 主表 codex（表头+示例行，冻结首行、设置列宽）
# 2) 字典表 dict（槽位/稀有度/布尔选项），并添加下拉校验到主表对应列
# 3) 说明表 readme（字段含义与填写规范）
# 生成后的文件保存在 /mnt/data/equipment_codex_template.xlsx

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

# ========= 1. 新建工作簿/工作表 =========
wb = Workbook()
ws = wb.active
ws.title = "codex"  # 主数据表

# ========= 2. 写入表头 =========
headers = [
    "equip_key","name","slot","rarity","star_max","level_requirement","color_hex",
    "icon","description","sort_order","enabled","release_version",
    "attr_attack","attr_atk_speed_pct","attr_crit_rate_pct","attr_crit_dmg_pct","attr_hit_pct","attr_penetration",
    "attr_metal","attr_wood","attr_water","attr_fire","attr_earth","attr_chaos",
    "tags","dungeon_keys"
]
ws.append(headers)

# 表头样式：加粗、垂直居中
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(vertical="center")

# 冻结首行
ws.freeze_panes = "A2"

# 设置列宽（便于阅读）
widths = [
    20, 18, 14, 12, 10, 18, 12,
    28, 40, 10, 10, 14,
    12, 16, 16, 16, 12, 14,
    12, 12, 12, 12, 12, 12,
    28, 22
]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64+i)].width = w  # A=65

# ========= 3. 示例行（第2行，可根据需要继续添加更多示例） =========
demo_row_1 = [
    "glove_taiji", "太极手套", "glove", "COMMON", 5, 1, "#9ca3af",
    "icons/gloves/taiji.png", "入门拳法手套，蕴含太极之意。", 0, True, "1.0.0",
    10, 0, 0, 0, 0, 0,
    0, 0, 0, 0, 0, 0,
    "入门,拳套,太极", "equip"
]
ws.append(demo_row_1)

# 可选：再添加一个示例（不同槽位/稀有度）
demo_row_2 = [
    "cloak_shadow", "暗影斗篷", "cloak", "LEGENDARY", 5, 30, "",  # color_hex 留空，DB 可按稀有度自动补色
    "icons/cloak/shadow.png", "潜行者专属斗篷，吸收微光。", 10, True, "1.1.0",
    0, 5, 10, 50, 0, 0,
    0, 0, 0, 0, 0, 10,
    "斗篷,暗影", "equip"
]
ws.append(demo_row_2)

# ========= 4. 字典表（槽位/稀有度/布尔）并添加下拉校验 =========
dict_ws = wb.create_sheet("dict")
dict_ws.append(["slot", "rarity", "bool"])
slot_values = ["cloak","wrist","glove","armor","belt","pants","shoes","ring","necklace","shoulder"]
rarity_values = ["COMMON","UNCOMMON","RARE","EPIC","LEGENDARY","MYTHIC"]
bool_values = ["TRUE","FALSE"]

# 填充字典表数据
max_len = max(len(slot_values), len(rarity_values), len(bool_values))
for i in range(max_len):
    row = []
    row.append(slot_values[i] if i < len(slot_values) else None)
    row.append(rarity_values[i] if i < len(rarity_values) else None)
    row.append(bool_values[i] if i < len(bool_values) else None)
    dict_ws.append(row)

# 添加数据有效性（下拉）到主表：
# slot -> C 列（C2:C2000）
dv_slot = DataValidation(type="list", formula1="=dict!$A$2:$A$11", allow_blank=False, showErrorMessage=True)
dv_slot.error = "无效的槽位，请从下拉列表选择"
ws.add_data_validation(dv_slot)
dv_slot.add(f"C2:C2000")

# rarity -> D 列（D2:D2000）
dv_rarity = DataValidation(type="list", formula1="=dict!$B$2:$B$7", allow_blank=False, showErrorMessage=True)
dv_rarity.error = "无效的稀有度，请从下拉列表选择"
ws.add_data_validation(dv_rarity)
dv_rarity.add(f"D2:D2000")

# enabled -> K 列（K2:K2000）
dv_bool = DataValidation(type="list", formula1="=dict!$C$2:$C$3", allow_blank=False, showErrorMessage=True)
dv_bool.error = "enabled 只能为 TRUE 或 FALSE"
ws.add_data_validation(dv_bool)
dv_bool.add(f"K2:K2000")

# 可选：隐藏字典表，避免误改
dict_ws.sheet_state = "hidden"

# ========= 5. 说明表（可选） =========
readme = wb.create_sheet("readme")
lines = [
    "【填写说明】",
    "1. 必填列：equip_key / name / slot / rarity。slot 与 rarity 支持下拉选择。",
    "2. color_hex 可留空；若留空，后端数据库触发器会按稀有度自动填充颜色。",
    "3. tags 与 dungeon_keys 可用英文逗号、中文逗号或分号分隔；导入时会覆盖旧标签/来源。",
    "4. 数值字段为整数；enabled 为 TRUE/FALSE。",
    "5. 若数据库列为 level_req，后端实体已映射为 levelRequirement，无需修改 Excel 列名。",
    "6. 建议一次导入 1~5000 行；更多请分批处理。"
]
for i, text in enumerate(lines, start=1):
    readme.cell(row=i, column=1, value=text)
readme.column_dimensions["A"].width = 100

# ========= 6. 保存文件 =========
save_path = "/mnt/data/equipment_codex_template.xlsx"
wb.save(save_path)

save_path
